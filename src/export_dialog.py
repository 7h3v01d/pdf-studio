"""PDF export dialog.

Supports:
  • Microsoft Word (.docx) via pdf2docx
  • Microsoft Excel (.xlsx) via tabula-py + openpyxl
  • Page images via PyMuPDF + Pillow
"""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

from PyQt6.QtCore import QThread, pyqtSignal
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import (
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QSizePolicy,
    QSpinBox,
    QVBoxLayout,
)

from office_export_core import commit_ooxml_atomic

from image_export_core import (
    FORMAT_EXTENSIONS,
    QUALITY_FORMATS,
    TRANSPARENCY_FORMATS,
    ImageExportOptions,
    build_output_paths,
    export_pdf_pages,
    resolve_page_indices,
)


def _check_docx_deps() -> tuple[bool, str]:
    try:
        import pdf2docx  # noqa: F401
    except Exception as exc:
        return False, (
            "Word export needs the 'pdf2docx' package and its dependencies "
            "(including opencv / cv2).\n\n"
            f"Importing it failed with:\n{type(exc).__name__}: {exc}\n\n"
            "From source, run:  pip install pdf2docx\n"
            "In a built .exe, this usually means a dependency of pdf2docx "
            "wasn't bundled — rebuild with the current spec."
        )
    return True, ""


def _check_xlsx_deps() -> tuple[bool, str]:
    missing = []
    for pkg, install in [
        ("tabula", "tabula-py"),
        ("openpyxl", "openpyxl"),
        ("pandas", "pandas"),
    ]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(f"{pkg}  (pip install {install})")
    if missing:
        return False, "Missing:\n• " + "\n• ".join(missing)
    return True, ""


def _check_image_deps() -> tuple[bool, str]:
    try:
        import fitz  # noqa: F401
        from PIL import Image  # noqa: F401
    except Exception as exc:
        return False, (
            "Image export requires PyMuPDF and Pillow.\n\n"
            f"Importing them failed with:\n{type(exc).__name__}: {exc}"
        )
    return True, ""


class DocxWorker(QThread):
    progress = pyqtSignal(int, str)
    result = pyqtSignal(str)
    error = pyqtSignal(str)
    cancelled = pyqtSignal()

    def __init__(self, pdf_path: str, out_path: str, start_page: int, end_page: int):
        super().__init__()
        self.pdf_path = pdf_path
        self.out_path = out_path
        self.start_page = start_page
        self.end_page = end_page

    def run(self):
        try:
            from pdf2docx import Converter

            if self.isInterruptionRequested():
                raise InterruptedError
            self.progress.emit(5, "Initialising converter...")

            def _produce(staged_path: str):
                converter = Converter(self.pdf_path)
                try:
                    self.progress.emit(15, "Converting pages...")
                    converter.convert(
                        staged_path,
                        start=self.start_page,
                        end=self.end_page + 1,
                    )
                finally:
                    converter.close()
                if self.isInterruptionRequested():
                    raise InterruptedError

            output = commit_ooxml_atomic(self.out_path, "docx", _produce)
            if self.isInterruptionRequested():
                # Cancellation requested after commit is treated as success:
                # the validated destination already exists and must not be deleted.
                self.progress.emit(100, "Done")
            else:
                self.progress.emit(100, "Done")
            self.result.emit(output)
        except InterruptedError:
            self.cancelled.emit()
        except Exception as exc:
            self.error.emit(f"{type(exc).__name__}: {exc}")


class XlsxWorker(QThread):
    progress = pyqtSignal(int, str)
    result = pyqtSignal(str)
    error = pyqtSignal(str)
    cancelled = pyqtSignal()

    def __init__(self, pdf_path: str, out_path: str, all_pages: bool, page_idx: int):
        super().__init__()
        self.pdf_path = pdf_path
        self.out_path = out_path
        self.all_pages = all_pages
        self.page_idx = page_idx

    def run(self):
        try:
            import openpyxl
            import tabula
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            if self.isInterruptionRequested():
                raise InterruptedError
            self.progress.emit(5, "Scanning for tables...")
            pages = "all" if self.all_pages else (self.page_idx + 1)
            dataframes = tabula.read_pdf(
                self.pdf_path,
                pages=pages,
                multiple_tables=True,
                silent=True,
            )
            if self.isInterruptionRequested():
                raise InterruptedError
            if not dataframes:
                self.error.emit(
                    "No tables detected in the selected page(s).\n\n"
                    "Excel export extracts structured tables only. "
                    "Use Word or image export for other content."
                )
                return

            self.progress.emit(40, f"Found {len(dataframes)} table(s). Writing...")

            def _produce(staged_path: str):
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)
                header_fill = PatternFill("solid", fgColor="2563EB")
                header_font = Font(color="FFFFFF", bold=True, size=10)
                header_align = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                thin_side = Side(style="thin", color="CCCCCC")
                cell_border = Border(
                    left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side,
                )
                alt_fill = PatternFill("solid", fgColor="EFF6FF")

                for table_index, dataframe in enumerate(dataframes):
                    if self.isInterruptionRequested():
                        raise InterruptedError
                    dataframe = dataframe.dropna(how="all").fillna("")
                    sheet_name = f"Table {table_index + 1}"
                    worksheet = workbook.create_sheet(title=sheet_name)
                    for column_index, column_name in enumerate(
                        dataframe.columns, start=1
                    ):
                        cell = worksheet.cell(
                            row=1, column=column_index, value=str(column_name)
                        )
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = cell_border
                    worksheet.row_dimensions[1].height = 22

                    for row_index, row in enumerate(
                        dataframe.itertuples(index=False), start=2
                    ):
                        for column_index, value in enumerate(row, start=1):
                            cell = worksheet.cell(
                                row=row_index,
                                column=column_index,
                                value=str(value) if value != "" else "",
                            )
                            cell.border = cell_border
                            if row_index % 2 == 0:
                                cell.fill = alt_fill

                    for column in worksheet.columns:
                        maximum = max(len(str(cell.value or "")) for cell in column)
                        worksheet.column_dimensions[
                            column[0].column_letter
                        ].width = min(maximum + 4, 50)
                    worksheet.freeze_panes = "A2"
                    percentage = 40 + int(
                        55 * (table_index + 1) / len(dataframes)
                    )
                    self.progress.emit(percentage, f"Written {sheet_name}...")

                if self.isInterruptionRequested():
                    raise InterruptedError
                workbook.save(staged_path)
                if self.isInterruptionRequested():
                    raise InterruptedError

            output = commit_ooxml_atomic(self.out_path, "xlsx", _produce)
            self.progress.emit(100, "Done")
            self.result.emit(output)
        except InterruptedError:
            self.cancelled.emit()
        except Exception as exc:
            self.error.emit(f"{type(exc).__name__}: {exc}")


class ImageWorker(QThread):
    progress = pyqtSignal(int, str)
    result = pyqtSignal(object)
    error = pyqtSignal(str)
    cancelled = pyqtSignal()

    def __init__(
        self,
        pdf_path: str,
        page_indices: list[int],
        output_paths: list[Path],
        options: ImageExportOptions,
    ):
        super().__init__()
        self.pdf_path = pdf_path
        self.page_indices = page_indices
        self.output_paths = output_paths
        self.options = options

    def run(self):
        try:
            outputs = export_pdf_pages(
                self.pdf_path,
                self.page_indices,
                self.output_paths,
                self.options,
                progress=lambda pct, msg: self.progress.emit(pct, msg),
                cancelled=self.isInterruptionRequested,
            )
            self.result.emit(outputs)
        except InterruptedError:
            self.cancelled.emit()
        except Exception as exc:
            self.error.emit(f"{type(exc).__name__}: {exc}")


class ExportDialog(QDialog):
    """Export PDF content to Word, Excel, or page image files."""

    FORMAT_DOCX = 0
    FORMAT_XLSX = 1
    FORMAT_IMAGE = 2

    def __init__(self, pdf_path: str, total_pages: int, current_page: int, parent=None):
        super().__init__(parent)
        self.pdf_path = pdf_path
        self.total_pages = total_pages
        self.current_page = current_page
        self._worker: QThread | None = None
        self._worker_result = None
        self._last_open_target: str | None = None
        self._close_when_worker_stops = False

        self.setWindowTitle("Export PDF As…")
        self.setMinimumWidth(520)
        self.setModal(True)
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setSpacing(12)
        root.setContentsMargins(16, 16, 16, 16)

        title = QLabel("📤  Export PDF As…")
        font = QFont()
        font.setPointSize(13)
        font.setBold(True)
        title.setFont(font)
        root.addWidget(title)

        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        root.addWidget(separator)

        format_box = QGroupBox("Export format")
        format_layout = QVBoxLayout(format_box)
        self._fmt_group = QButtonGroup(self)
        self._rb_docx = QRadioButton(
            "📝  Microsoft Word (.docx)\n    Preserves editable text and page layout where possible."
        )
        self._rb_xlsx = QRadioButton(
            "📊  Microsoft Excel (.xlsx)\n    Extracts structured tables into spreadsheet sheets."
        )
        self._rb_image = QRadioButton(
            "🖼️  Image files\n    Renders selected PDF pages as PNG, JPEG, WebP, TIFF, BMP, or GIF."
        )
        self._rb_docx.setChecked(True)
        for button, identifier in (
            (self._rb_docx, self.FORMAT_DOCX),
            (self._rb_xlsx, self.FORMAT_XLSX),
            (self._rb_image, self.FORMAT_IMAGE),
        ):
            self._fmt_group.addButton(button, identifier)
            format_layout.addWidget(button)
        root.addWidget(format_box)

        self._scope_box = QGroupBox("Pages")
        scope_layout = QVBoxLayout(self._scope_box)
        self._scope_group = QButtonGroup(self)
        self._rb_all_pages = QRadioButton(f"All pages ({self.total_pages})")
        self._rb_cur_page = QRadioButton(f"Current page (page {self.current_page + 1})")
        self._rb_pg_range = QRadioButton("Page range:")
        self._rb_all_pages.setChecked(True)
        for button, identifier in (
            (self._rb_all_pages, 0),
            (self._rb_cur_page, 1),
            (self._rb_pg_range, 2),
        ):
            self._scope_group.addButton(button, identifier)
            scope_layout.addWidget(button)
        range_row = QHBoxLayout()
        range_row.setContentsMargins(20, 0, 0, 0)
        self._page_range_input = QLineEdit()
        self._page_range_input.setPlaceholderText("e.g. 1-5, 8, 12-15")
        self._page_range_input.setEnabled(False)
        self._page_range_input.setFixedWidth(220)
        range_row.addWidget(self._page_range_input)
        range_row.addStretch()
        scope_layout.addLayout(range_row)
        root.addWidget(self._scope_box)
        self._rb_pg_range.toggled.connect(self._page_range_input.setEnabled)

        self._image_options = QGroupBox("Image options")
        image_layout = QVBoxLayout(self._image_options)

        format_row = QHBoxLayout()
        format_row.addWidget(QLabel("Image format:"))
        self._image_format = QComboBox()
        self._image_format.addItems(["PNG", "JPEG", "WebP", "TIFF", "BMP", "GIF"])
        format_row.addWidget(self._image_format)
        format_row.addStretch()
        image_layout.addLayout(format_row)

        resolution_row = QHBoxLayout()
        resolution_row.addWidget(QLabel("Resolution:"))
        self._image_dpi = QSpinBox()
        self._image_dpi.setRange(36, 1200)
        self._image_dpi.setValue(300)
        self._image_dpi.setSuffix(" DPI")
        self._image_dpi.setToolTip(
            "300 DPI is suitable for printing. 150 DPI is usually enough for screens."
        )
        resolution_row.addWidget(self._image_dpi)
        resolution_row.addStretch()
        image_layout.addLayout(resolution_row)

        quality_row = QHBoxLayout()
        self._quality_label = QLabel("Quality:")
        quality_row.addWidget(self._quality_label)
        self._image_quality = QSpinBox()
        self._image_quality.setRange(1, 100)
        self._image_quality.setValue(92)
        self._image_quality.setSuffix("%")
        quality_row.addWidget(self._image_quality)
        quality_row.addStretch()
        image_layout.addLayout(quality_row)

        self._transparent = QCheckBox("Transparent page background where supported")
        self._transparent.setChecked(False)
        image_layout.addWidget(self._transparent)

        self._image_note = QLabel(
            "Multi-page exports create one numbered image per page. GIF export is static, not animated."
        )
        self._image_note.setWordWrap(True)
        self._image_note.setStyleSheet("color:#64748b; font-size:11px;")
        image_layout.addWidget(self._image_note)
        root.addWidget(self._image_options)

        self._cb_open = QCheckBox("Open exported file or folder when finished")
        self._cb_open.setChecked(True)
        root.addWidget(self._cb_open)

        self._progress = QProgressBar()
        self._progress.setVisible(False)
        self._progress.setRange(0, 100)
        root.addWidget(self._progress)

        self._progress_label = QLabel("")
        self._progress_label.setVisible(False)
        self._progress_label.setStyleSheet("color:#555; font-size:11px;")
        root.addWidget(self._progress_label)

        button_row = QHBoxLayout()
        self._btn_export = QPushButton("Export")
        self._btn_export.setDefault(True)
        self._btn_export.setStyleSheet(
            "background:#3a7bd5; color:white; font-weight:bold;"
            "border-radius:4px; padding:6px 18px;"
        )
        self._btn_close = QPushButton("Cancel")
        button_row.addStretch()
        button_row.addWidget(self._btn_close)
        button_row.addWidget(self._btn_export)
        root.addLayout(button_row)

        self._fmt_group.idToggled.connect(self._on_format_changed)
        self._image_format.currentTextChanged.connect(self._on_image_format_changed)
        self._btn_export.clicked.connect(self._do_export)
        self._btn_close.clicked.connect(self._on_close)
        self._on_format_changed(self._fmt_group.checkedId(), True)
        self._on_image_format_changed(self._image_format.currentText())

    def select_format(self, fmt: str) -> None:
        key = fmt.lower().strip()
        if key == "xlsx":
            self._rb_xlsx.setChecked(True)
        elif key in {"image", "images", "png", "jpg", "jpeg", "webp", "tiff", "bmp", "gif"}:
            self._rb_image.setChecked(True)
            label_map = {
                "jpg": "JPEG",
                "jpeg": "JPEG",
                "webp": "WebP",
                "tiff": "TIFF",
                "bmp": "BMP",
                "gif": "GIF",
                "png": "PNG",
            }
            desired = label_map.get(key)
            if desired:
                index = self._image_format.findText(desired)
                if index >= 0:
                    self._image_format.setCurrentIndex(index)
        else:
            self._rb_docx.setChecked(True)

    def _on_format_changed(self, _identifier: int, _checked: bool):
        selected = self._fmt_group.checkedId()
        is_image = selected == self.FORMAT_IMAGE
        is_xlsx = selected == self.FORMAT_XLSX
        self._image_options.setVisible(is_image)
        self._scope_box.setEnabled(not is_xlsx)
        if is_xlsx:
            self._rb_all_pages.setChecked(True)
        self.adjustSize()

    def _on_image_format_changed(self, label: str):
        fmt = label.upper()
        quality_enabled = fmt in QUALITY_FORMATS
        self._quality_label.setEnabled(quality_enabled)
        self._image_quality.setEnabled(quality_enabled)
        transparency_enabled = fmt in TRANSPARENCY_FORMATS
        self._transparent.setEnabled(transparency_enabled)
        if not transparency_enabled:
            self._transparent.setChecked(False)

    def _do_export(self):
        selected = self._fmt_group.checkedId()
        if selected == self.FORMAT_XLSX:
            ok, message = _check_xlsx_deps()
        elif selected == self.FORMAT_IMAGE:
            ok, message = _check_image_deps()
        else:
            ok, message = _check_docx_deps()
        if not ok:
            QMessageBox.critical(self, "Missing Dependencies", message)
            return

        if selected == self.FORMAT_IMAGE:
            self._start_image_export()
            return

        extension = ".xlsx" if selected == self.FORMAT_XLSX else ".docx"
        description = (
            "Excel Files (*.xlsx)"
            if selected == self.FORMAT_XLSX
            else "Word Documents (*.docx)"
        )
        base = os.path.splitext(self.pdf_path)[0]
        out_path, _ = QFileDialog.getSaveFileName(
            self, "Save Export As", base + extension, description
        )
        if not out_path:
            return

        self._set_running_state()
        if selected == self.FORMAT_XLSX:
            self._worker = XlsxWorker(
                self.pdf_path,
                out_path,
                all_pages=True,
                page_idx=self.current_page,
            )
        else:
            page_indices = self._resolve_pages()
            if page_indices is None:
                self._restore_idle_state()
                return
            # pdf2docx accepts a contiguous interval only.
            expected = list(range(min(page_indices), max(page_indices) + 1))
            if page_indices != expected:
                self._restore_idle_state()
                QMessageBox.warning(
                    self,
                    "Word Page Range",
                    "Word export currently requires a continuous page range.",
                )
                return
            self._worker = DocxWorker(
                self.pdf_path,
                out_path,
                min(page_indices),
                max(page_indices),
            )
        self._last_open_target = out_path
        self._connect_worker(self._worker)
        self._worker.start()

    def _start_image_export(self):
        page_indices = self._resolve_pages()
        if page_indices is None:
            return

        fmt_label = self._image_format.currentText()
        fmt = fmt_label.upper()
        extension = FORMAT_EXTENSIONS[fmt]
        base = Path(self.pdf_path)

        if len(page_indices) == 1:
            page_number = page_indices[0] + 1
            default = base.with_name(f"{base.stem}_page_{page_number}{extension}")
            filters = {
                "PNG": "PNG Images (*.png)",
                "JPEG": "JPEG Images (*.jpg *.jpeg)",
                "WEBP": "WebP Images (*.webp)",
                "TIFF": "TIFF Images (*.tif *.tiff)",
                "BMP": "Bitmap Images (*.bmp)",
                "GIF": "GIF Images (*.gif)",
            }
            destination, _ = QFileDialog.getSaveFileName(
                self, "Save Page Image", str(default), filters[fmt]
            )
            if not destination:
                return
        else:
            destination = QFileDialog.getExistingDirectory(
                self,
                "Choose Folder for Exported Page Images",
                str(base.parent),
            )
            if not destination:
                return

        options = ImageExportOptions(
            image_format=fmt,
            dpi=self._image_dpi.value(),
            quality=self._image_quality.value(),
            transparent_background=self._transparent.isChecked(),
        ).normalised()
        output_paths = build_output_paths(
            self.pdf_path,
            destination,
            page_indices,
            options.image_format,
            self.total_pages,
        )
        existing = [path for path in output_paths if path.exists()]
        if existing:
            answer = QMessageBox.question(
                self,
                "Replace Existing Images?",
                f"{len(existing)} destination image(s) already exist. Replace them?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if answer != QMessageBox.StandardButton.Yes:
                return

        self._set_running_state()
        self._last_open_target = (
            str(output_paths[0]) if len(output_paths) == 1 else str(Path(destination))
        )
        self._worker = ImageWorker(
            self.pdf_path,
            page_indices,
            output_paths,
            options,
        )
        self._worker.progress.connect(self._on_progress)
        self._worker.result.connect(
            lambda outputs: setattr(self, "_worker_result", ("images", outputs))
        )
        self._worker.error.connect(
            lambda message: setattr(self, "_worker_result", ("error", message))
        )
        self._worker.cancelled.connect(
            lambda: setattr(self, "_worker_result", ("cancelled", None))
        )
        self._worker.finished.connect(self._on_worker_thread_finished)
        self._worker.start()

    def _connect_worker(self, worker: QThread):
        worker.progress.connect(self._on_progress)
        worker.result.connect(
            lambda output: setattr(self, "_worker_result", ("success", output))
        )
        worker.error.connect(
            lambda message: setattr(self, "_worker_result", ("error", message))
        )
        worker.cancelled.connect(
            lambda: setattr(self, "_worker_result", ("cancelled", None))
        )
        worker.finished.connect(self._on_worker_thread_finished)

    def _set_running_state(self):
        self._close_when_worker_stops = False
        self._btn_export.setEnabled(False)
        self._btn_close.setText("Stop")
        self._progress.setValue(0)
        self._progress.setVisible(True)
        self._progress_label.setVisible(True)
        self._progress_label.setText("Starting…")

    def _restore_idle_state(self):
        self._btn_export.setEnabled(True)
        self._btn_close.setText("Cancel")
        self._progress.setVisible(False)
        self._progress_label.setVisible(False)

    def _on_progress(self, percentage: int, message: str):
        self._progress.setValue(percentage)
        self._progress_label.setText(message)

    def _finish_success(self, message: str):
        self._progress.setValue(100)
        self._progress_label.setText(message)
        self._btn_close.setText("Close")
        try:
            self._btn_close.clicked.disconnect()
        except TypeError:
            pass
        self._btn_close.clicked.connect(self.accept)
        self._btn_export.setEnabled(False)
        if self._cb_open.isChecked() and self._last_open_target:
            self._open_file(self._last_open_target)

    def _on_finished(self, _out_path: str):
        self._finish_success("Export complete ✓")

    def _on_images_finished(self, output_paths: list[Path]):
        count = len(output_paths)
        noun = "image" if count == 1 else "images"
        self._finish_success(f"Exported {count} {noun} ✓")

    def _on_cancelled(self):
        self._restore_idle_state()
        self._progress_label.setText("Export cancelled")

    def _on_error(self, message: str):
        self._restore_idle_state()
        QMessageBox.critical(self, "Export Error", message)

    def _on_worker_thread_finished(self):
        result = self._worker_result
        closing = self._close_when_worker_stops
        self._release_worker()
        if closing:
            QDialog.reject(self)
            return
        if result is None:
            self._on_error("The export worker stopped without reporting a result.")
            return
        kind, payload = result
        if kind == "success":
            self._on_finished(payload)
        elif kind == "images":
            self._on_images_finished(payload)
        elif kind == "cancelled":
            self._on_cancelled()
        else:
            self._on_error(payload)

    def _release_worker(self):
        worker = self._worker
        self._worker = None
        self._worker_result = None
        self._close_when_worker_stops = False
        self._btn_close.setEnabled(True)
        if worker is not None:
            worker.deleteLater()


    def _request_worker_cancel(self) -> bool:
        if self._worker is None or not self._worker.isRunning():
            return False
        reply = QMessageBox.question(
            self,
            "Cancel Export?",
            "Export is still running. Cancel it and close this window after "
            "the worker has stopped safely?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return True
        self._close_when_worker_stops = True
        self._worker.requestInterruption()
        self._btn_close.setEnabled(False)
        self._progress_label.setText("Cancelling safely...")
        return True

    def _on_close(self):
        if self._request_worker_cancel():
            return
        self.reject()

    def reject(self):
        if self._request_worker_cancel():
            return
        QDialog.reject(self)

    def closeEvent(self, event):
        if self._request_worker_cancel():
            event.ignore()
            return
        event.accept()

    def _resolve_pages(self) -> list[int] | None:
        try:
            return resolve_page_indices(
                self._scope_group.checkedId(),
                self._page_range_input.text(),
                self.total_pages,
                self.current_page,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Page Range", str(exc))
            return None

    @staticmethod
    def _open_file(path: str):
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", path], check=False)
            else:
                subprocess.run(["xdg-open", path], check=False)
        except Exception:
            pass
